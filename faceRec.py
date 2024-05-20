import cv2
import face_recognition
import os
import datetime
import openpyxl
import numpy as np

# Load the reference face images
reference_faces_dir = 'reference_faces'
reference_faces = []
reference_face_names = []

for filename in os.listdir(reference_faces_dir):
    img = face_recognition.load_image_file(os.path.join(reference_faces_dir, filename))
    reference_faces.append(face_recognition.face_encodings(img)[0])
    reference_face_names.append(os.path.splitext(filename)[0])

# Initialize the Excel file
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Attendance'

ws['A1'] = 'Name'
ws['B1'] = 'Timestamp'

# Initialize the camera
cap = cv2.VideoCapture(0)

stop_capturing = False
frame_count = 0
while True:
    # Capture a frame from the camera
    ret, frame = cap.read()

    # Convert the frame to grayscale
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Detect faces in the frame
    faces = face_recognition.face_locations(gray)
    encodings = face_recognition.face_encodings(frame, faces)

    # Loop through each face in the frame
    for encoding in encodings:
        # Find the best match for the face
        distances = face_recognition.face_distance(reference_faces, encoding)
        best_match_index = np.argmin(distances)

        # Check if the minimum distance is below a certain threshold (e.g., 0.6)
        if distances[best_match_index] < 0.6:
            name = reference_face_names[best_match_index]
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"Best match for the face: {name} at {timestamp}")

            # Write the name and timestamp to the Excel file
            ws.append([name, timestamp])
            wb.save('attendance.xlsx')
        else:
            print("Person not present in the reference dataset")
    frame_count += 1
    # Display the frame
    cv2.imshow('frame', frame)
    

    frame_count += 1

    # Check if the frame counter has reached the desired number of frames (e.g., 2)
    if frame_count >= 2:
        stop_capturing = True
        
    # Exit the loop if the user presses the 'q' key
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the camera and close any open windows
cap.release()
cv2.destroyAllWindows()